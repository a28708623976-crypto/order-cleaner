from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import FINAL_COLUMNS


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "清洗结果") -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        indexes = {name: i + 1 for i, name in enumerate(FINAL_COLUMNS)}
        for column in ["主订单编号", "子订单编号", "快递信息 (单号)"]:
            idx = indexes[column]
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=idx)
                if cell.value not in (None, ""):
                    cell.value = str(cell.value)
                cell.number_format = "@"

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        widths = {
            "主订单编号": 22, "子订单编号": 22, "选购商品": 48, "产品": 16,
            "重量规格": 12, "果型": 12, "供应规格": 16, "商品数量": 10,
            "商品金额": 12, "订单提交时间": 14, "订单完成时间": 14,
            "订单状态": 14, "售后状态": 14, "发货时间": 14,
            "收货地址": 45, "快递信息 (单号)": 24, "快递公司": 16,
        }
        for name, width in widths.items():
            worksheet.column_dimensions[get_column_letter(indexes[name])].width = width

    return buffer.getvalue()
